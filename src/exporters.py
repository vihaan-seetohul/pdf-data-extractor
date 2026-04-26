"""
Multi-format exporters for extracted invoice data.

Three output modes:
1. JSON - structured nested data, one file per invoice (or one combined file)
2. CSV - flat tabular, two layouts available:
   - "summary": one row per invoice (totals only)
   - "line_items": one row per line item (with invoice metadata repeated)
3. Excel - multi-sheet workbook with summary + line items + raw extraction
"""

import csv
import json
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# JSON EXPORT
# ============================================================

def export_json(extraction_results, output_path, indent=2):
    """
    Export extracted data to JSON.

    Args:
        extraction_results: list of dicts from extract_invoice()
        output_path: where to write the JSON file
        indent: JSON indentation level
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    payload = {
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "invoice_count": len(extraction_results),
        "invoices": extraction_results,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=indent, default=str)

    return output_path


# ============================================================
# CSV EXPORT
# ============================================================

SUMMARY_FIELDS = [
    "source_file", "invoice_number", "issue_date", "due_date",
    "vendor_name", "customer_name", "subtotal", "tax", "total",
    "line_item_count", "extraction_method",
]

LINE_ITEM_FIELDS = [
    "source_file", "invoice_number", "issue_date", "vendor_name",
    "customer_name", "line_number", "description", "sku",
    "quantity", "unit_price", "line_total",
]


def export_csv_summary(extraction_results, output_path):
    """One row per invoice."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    for r in extraction_results:
        row = {field: r.get(field) for field in SUMMARY_FIELDS if field != "line_item_count"}
        row["line_item_count"] = len(r.get("line_items") or [])
        rows.append(row)

    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    return output_path


def export_csv_line_items(extraction_results, output_path):
    """One row per line item, with invoice metadata repeated."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    for r in extraction_results:
        invoice_meta = {
            "source_file": r.get("source_file"),
            "invoice_number": r.get("invoice_number"),
            "issue_date": r.get("issue_date"),
            "vendor_name": r.get("vendor_name"),
            "customer_name": r.get("customer_name"),
        }
        items = r.get("line_items") or []
        if not items:
            # Still record the invoice with empty line item fields
            row = dict(invoice_meta)
            row["line_number"] = None
            row["description"] = None
            row["sku"] = None
            row["quantity"] = None
            row["unit_price"] = None
            row["line_total"] = None
            rows.append(row)
        else:
            for idx, item in enumerate(items, start=1):
                row = dict(invoice_meta)
                row["line_number"] = idx
                row["description"] = item.get("description")
                row["sku"] = item.get("sku")
                row["quantity"] = item.get("quantity")
                row["unit_price"] = item.get("unit_price")
                row["line_total"] = item.get("line_total")
                rows.append(row)

    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=LINE_ITEM_FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    return output_path


# ============================================================
# EXCEL EXPORT
# ============================================================

# Styling
HEADER_FILL = PatternFill(start_color="0A1929", end_color="0A1929", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALT_ROW_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _autosize_columns(worksheet, max_width=50):
    """Auto-fit column widths based on content."""
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        adjusted_width = min(max_length + 2, max_width)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max(adjusted_width, 10)


def _style_header_row(worksheet, num_cols):
    for col in range(1, num_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    worksheet.row_dimensions[1].height = 28


def _style_data_rows(worksheet, num_rows, num_cols):
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL


def export_excel(extraction_results, output_path):
    """
    Export to a multi-sheet Excel workbook:
    - Sheet 1 'Summary': one row per invoice
    - Sheet 2 'Line Items': one row per line item
    - Sheet 3 'Extraction Notes': metadata about extraction quality
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # === Sheet 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = [
        "Source File", "Invoice #", "Issue Date", "Due Date",
        "Vendor", "Customer", "Subtotal", "Tax", "Total",
        "# Line Items", "Method",
    ]
    ws_summary.append(summary_headers)

    for r in extraction_results:
        ws_summary.append([
            r.get("source_file"),
            r.get("invoice_number"),
            r.get("issue_date"),
            r.get("due_date"),
            r.get("vendor_name"),
            r.get("customer_name"),
            r.get("subtotal"),
            r.get("tax"),
            r.get("total"),
            len(r.get("line_items") or []),
            r.get("extraction_method"),
        ])

    _style_header_row(ws_summary, len(summary_headers))
    _style_data_rows(ws_summary, len(extraction_results), len(summary_headers))

    # Format currency columns (subtotal, tax, total are columns 7, 8, 9)
    for row in range(2, len(extraction_results) + 2):
        for col in [7, 8, 9]:
            cell = ws_summary.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = '"$"#,##0.00'

    _autosize_columns(ws_summary)

    # === Sheet 2: Line Items ===
    ws_items = wb.create_sheet("Line Items")
    line_headers = [
        "Source File", "Invoice #", "Issue Date", "Vendor", "Customer",
        "Line #", "Description", "SKU", "Qty", "Unit Price", "Line Total",
    ]
    ws_items.append(line_headers)

    line_count = 0
    for r in extraction_results:
        items = r.get("line_items") or []
        for idx, item in enumerate(items, start=1):
            ws_items.append([
                r.get("source_file"),
                r.get("invoice_number"),
                r.get("issue_date"),
                r.get("vendor_name"),
                r.get("customer_name"),
                idx,
                item.get("description"),
                item.get("sku"),
                item.get("quantity"),
                item.get("unit_price"),
                item.get("line_total"),
            ])
            line_count += 1

    _style_header_row(ws_items, len(line_headers))
    _style_data_rows(ws_items, line_count, len(line_headers))

    # Format currency columns (unit_price, line_total are columns 10, 11)
    for row in range(2, line_count + 2):
        for col in [10, 11]:
            cell = ws_items.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = '"$"#,##0.00'

    _autosize_columns(ws_items)

    # === Sheet 3: Extraction Notes ===
    ws_notes = wb.create_sheet("Extraction Notes")
    notes_headers = ["Source File", "Method", "Raw Text Length", "Has Line Items", "Has Customer"]
    ws_notes.append(notes_headers)

    for r in extraction_results:
        ws_notes.append([
            r.get("source_file"),
            r.get("extraction_method"),
            r.get("raw_text_length"),
            "Yes" if r.get("line_items") else "No",
            "Yes" if r.get("customer_name") else "No",
        ])

    _style_header_row(ws_notes, len(notes_headers))
    _style_data_rows(ws_notes, len(extraction_results), len(notes_headers))
    _autosize_columns(ws_notes)

    wb.save(output_path)
    return output_path


# ============================================================
# UNIFIED EXPORT INTERFACE
# ============================================================

def export_all(extraction_results, output_dir="output", base_name="invoices"):
    """
    Convenience function: export to all 4 formats at once.
    Returns dict mapping format name to output path.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    paths = {
        "json": export_json(extraction_results, output_dir / f"{base_name}.json"),
        "csv_summary": export_csv_summary(extraction_results, output_dir / f"{base_name}_summary.csv"),
        "csv_line_items": export_csv_line_items(extraction_results, output_dir / f"{base_name}_line_items.csv"),
        "excel": export_excel(extraction_results, output_dir / f"{base_name}.xlsx"),
    }
    return paths